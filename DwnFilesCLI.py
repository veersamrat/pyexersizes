import openpyxl as xl
import requests
import os
import concurrent.futures 
import argparse

parser = argparse.ArgumentParser(description="Candidate Recordings downloader for Eduswitch Excel File ")

parser.add_argument("-xlf","--f",help="The Path of Excel file from Eduswitch",required=True)
parser.add_argument("-dest","--d",help="The Path for recording to save",required=True)
parser.add_argument("-wrk","--w",help="The Concurrent Process Count ",default=4)
parser.add_argument("-lnlCol","--lc",help="The column number of video download url column, default column no is 14",default=14)
parser.add_argument("-fnCol","--fc",help="The column number of Enrollment no column, default column no is 1",default=1)

args = parser.parse_args()
#print(type(args))
videosPath = args.d+"\\"
xlPath = args.f
maxWorkers = args.w
lnkColumn = args.lc  
nameColumn = args.fc

#print(videosPath,'\n',xlPath,"\n",maxWorkers)

# Define variable to load the dataframe
dataframe = xl.load_workbook(xlPath)

# Define variable to read sheet
dataframe1 = dataframe.active

# names=list()
# urls=list()
urls_data = dict()
for row in dataframe1.iter_rows(1, dataframe1.max_row):
    if 'https' in row[14].value:
        urls_data.update({row[nameColumn].value:row[lnkColumn].value})
        # names.append(row[1].value)
        # urls.append(row[14].value)
        # print(row[14].value)

def download_recordings(key):
    name=key
    url=urls_data.get(key)
    try:
        filePath=videosPath + name + '.mp4'
        if os.path.isfile(filePath) == False:
            print(f'Started downloading {filePath}')
            response=requests.get(url)
            open(filePath,"wb").write(response.content)
            print(f'Finished downloading {filePath}')
    except:
        print(f"Got Error while downloading {name}, url:{url}")

if __name__ == '__main__':
    with concurrent.futures.ProcessPoolExecutor(max_workers=maxWorkers) as exector : 
        exector.map(download_recordings,urls_data)
    print("All videos downloaded")
   
        



