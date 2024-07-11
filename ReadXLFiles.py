import openpyxl as xl
import requests
import os
import concurrent.futures 

maxWorkers=4 # Maximum concurrent process required
videosPath='D:\\temp\\record\\videos\\' # Videos Path

# Define variable to load the dataframe
dataframe = xl.load_workbook("D:\\temp\\record\\23.3.2024_PG.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
# for row in range(0, dataframe1.max_row):
#     for col in dataframe1.iter_cols(1, dataframe1.max_column):
#         print(col[row].value)
names=list()
urls=list()
for row in dataframe1.iter_rows(1, dataframe1.max_row):
    if 'https' in row[14].value:
        names.append(row[1].value)
        urls.append(row[14].value)
        #print(row[14].value)

def download_recordings(url,name):
    try:
        filePath=videosPath+name+'.mp4'
        if os.path.isfile(filePath)==False:
            print(f'Started downloading {filePath}')
            response=requests.get(url)
            open(filePath,"wb").write(response.content)
            print(f'Finished downloading {filePath}')
    except:
        print(f"Got Error while downloading {name}, url:{url}")

if __name__ == '__main__':
    if len(names)==len(urls):
        with concurrent.futures.ProcessPoolExecutor(max_workers=maxWorkers) as exector : 
            exector.map(download_recordings,urls,names)
        print("All videos downloaded")
    else:
        print("Videos and Names counts are not matching")


# if __name__ == "__main__":
#     print("There are {} CPUs on this machine ".format(cpu_count()))
#     pool = Pool(cpu_count())
#     download_func = partial(download_recordings, names = names)
#     results = pool.map(download_recordings, urls,names)
#     pool.close()
#     pool.join()
    
    


    
    